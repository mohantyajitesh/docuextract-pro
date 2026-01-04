"""
DocuExtract Pro - Document Processing Engine
Core extraction logic for text, tables, signatures, and key-value pairs.
"""
import os
import json
import base64
import tempfile
import re
import time
from pathlib import Path
from typing import Optional, Dict, Any, List, Callable
from datetime import datetime

import cv2
import numpy as np
from PIL import Image

from langchain_ollama import ChatOllama
from langchain_core.messages import HumanMessage

from .config import (
    TEXT_MODEL, VISION_MODEL, OLLAMA_BASE_URL,
    MAX_TOKENS, DOC_TEXT_LIMIT, SIGNATURE_CONFIDENCE_THRESHOLD, OUTPUT_DIR
)
from .models import (
    SignatureResult, SignatureStatus, KeyValuePair, TableData,
    HumanReviewItem, ExtractionResult, JobStatus
)


# =============================================================================
# EASYOCR SETUP (Lazy initialization - M1 compatible)
# =============================================================================

_ocr_reader = None


def get_ocr_reader():
    """Lazy initialization of EasyOCR with MPS support for M1."""
    global _ocr_reader
    if _ocr_reader is None:
        import easyocr
        _ocr_reader = easyocr.Reader(
            ['en'],
            gpu=True,  # Uses MPS on M1 Macs
            verbose=False
        )
    return _ocr_reader


# =============================================================================
# SIGNATURE DETECTION (OpenCV)
# =============================================================================

def detect_signatures(image_path: str, threshold: float = SIGNATURE_CONFIDENCE_THRESHOLD) -> Dict:
    """
    Detect signatures in an image using OpenCV.
    Looks for ink-like marks in typical signature regions.
    """
    signatures = []
    review_items = []

    try:
        img = cv2.imread(image_path)
        if img is None:
            return {"signatures": [], "count": 0, "valid_count": 0, "human_review_items": []}

        height, width = img.shape[:2]
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        # Focus on bottom third (common signature location)
        signature_region = gray[int(height * 0.6):, :]

        # Adaptive threshold to find ink marks
        binary = cv2.adaptiveThreshold(
            signature_region, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY_INV, 11, 2
        )

        # Find contours
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        potential_signatures = []

        for contour in contours:
            area = cv2.contourArea(contour)
            if area < 500 or area > 50000:
                continue

            x, y, w, h = cv2.boundingRect(contour)
            aspect_ratio = w / h if h > 0 else 0

            if 1.5 < aspect_ratio < 10 and w > 50:
                roi = binary[y:y+h, x:x+w]
                ink_density = np.sum(roi > 0) / (w * h) if w * h > 0 else 0

                if 0.05 < ink_density < 0.5:
                    confidence = min(0.9, ink_density * 2 + 0.3)
                    potential_signatures.append({
                        "bbox": (x, y + int(height * 0.6), w, h),
                        "confidence": round(confidence, 3),
                        "ink_density": round(ink_density, 3)
                    })

        potential_signatures.sort(key=lambda s: s["confidence"], reverse=True)

        for i, sig in enumerate(potential_signatures[:3]):
            x, y, w, h = sig["bbox"]
            confidence = sig["confidence"]

            if confidence >= threshold:
                status = SignatureStatus.VALID
            elif confidence >= 0.4:
                status = SignatureStatus.NEEDS_REVIEW
            else:
                status = SignatureStatus.INVALID

            signature = SignatureResult(
                id=f"sig_{i+1}",
                confidence=confidence,
                location={
                    'left': round(x / width, 4),
                    'top': round(y / height, 4),
                    'width': round(w / width, 4),
                    'height': round(h / height, 4)
                },
                status=status
            )
            signatures.append(signature)

            if status == SignatureStatus.NEEDS_REVIEW:
                review_items.append(HumanReviewItem(
                    type='signature',
                    id=signature.id,
                    confidence=confidence,
                    reason=f"Confidence {confidence*100:.0f}% below {threshold*100:.0f}% threshold"
                ))

    except Exception as e:
        print(f"Signature detection error: {e}")

    return {
        'signatures': signatures,
        'count': len(signatures),
        'valid_count': sum(1 for s in signatures if s.status == SignatureStatus.VALID),
        'human_review_items': review_items
    }


# =============================================================================
# OCR WITH EASYOCR
# =============================================================================

def extract_with_easyocr(image_path: str) -> Dict:
    """Extract text using EasyOCR (M1 compatible with MPS support)."""
    reader = get_ocr_reader()
    result = reader.readtext(image_path)

    lines = []
    all_text_blocks = []

    for detection in result:
        bbox, text, confidence = detection
        all_text_blocks.append({
            'text': text,
            'confidence': confidence,
            'bbox': bbox
        })
        lines.append(text)

    # Simple key-value extraction
    key_values = []
    full_text = ' '.join(lines)
    kv_pattern = r'([A-Za-z][A-Za-z\s]{2,30}):\s*([^\n:]{1,100})'
    matches = re.findall(kv_pattern, full_text)

    for key, value in matches:
        key_values.append(KeyValuePair(
            key=key.strip(),
            value=value.strip(),
            confidence=0.8
        ))

    return {
        'text': '\n'.join(lines),
        'lines': lines,
        'key_values': key_values,
        'raw_blocks': all_text_blocks
    }


def extract_tables_with_img2table(image_path: str) -> List[TableData]:
    """Extract tables using img2table."""
    tables = []

    try:
        from img2table.document import Image as Img2TableImage
        from img2table.ocr import EasyOCR as Img2TableOCR

        ocr = Img2TableOCR(lang=['en'])
        doc = Img2TableImage(src=image_path)
        extracted_tables = doc.extract_tables(ocr=ocr)

        for idx, table in enumerate(extracted_tables):
            if table.df is not None:
                rows = [table.df.columns.tolist()] + table.df.values.tolist()
                tables.append(TableData(
                    id=f"table_{idx+1}",
                    rows=[[str(cell) for cell in row] for row in rows],
                    headers=table.df.columns.tolist()
                ))

    except Exception as e:
        print(f"img2table extraction error: {e}")
        tables = extract_tables_opencv_fallback(image_path)

    return tables


def extract_tables_opencv_fallback(image_path: str) -> List[TableData]:
    """Fallback table detection using OpenCV."""
    tables = []

    try:
        img = cv2.imread(image_path)
        if img is None:
            return tables

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))

        horizontal_lines = cv2.morphologyEx(gray, cv2.MORPH_OPEN, horizontal_kernel)
        vertical_lines = cv2.morphologyEx(gray, cv2.MORPH_OPEN, vertical_kernel)

        table_mask = cv2.addWeighted(horizontal_lines, 0.5, vertical_lines, 0.5, 0)
        _, table_mask = cv2.threshold(table_mask, 128, 255, cv2.THRESH_BINARY)

        contours, _ = cv2.findContours(table_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

        if len(contours) > 4:
            reader = get_ocr_reader()
            cells = []

            for contour in contours:
                x, y, w, h = cv2.boundingRect(contour)
                if w > 30 and h > 15:
                    cell_img = img[y:y+h, x:x+w]
                    temp_cell = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                    cv2.imwrite(temp_cell.name, cell_img)
                    result = reader.readtext(temp_cell.name)
                    os.unlink(temp_cell.name)
                    cell_text = ' '.join([det[1] for det in result]) if result else ''
                    cells.append({'x': x, 'y': y, 'w': w, 'h': h, 'text': cell_text})

            if cells:
                cells.sort(key=lambda c: (c['y'] // 20, c['x']))
                table_rows = []
                current_row = []
                last_y = -100

                for cell in cells:
                    if abs(cell['y'] - last_y) > 15:
                        if current_row:
                            table_rows.append([c['text'] for c in current_row])
                        current_row = [cell]
                        last_y = cell['y']
                    else:
                        current_row.append(cell)

                if current_row:
                    table_rows.append([c['text'] for c in current_row])

                if table_rows:
                    tables.append(TableData(
                        id="table_1",
                        rows=table_rows
                    ))

    except Exception as e:
        print(f"OpenCV table extraction error: {e}")

    return tables


# =============================================================================
# DOCUMENT LOADING
# =============================================================================

def load_with_docling(file_path: Path) -> str:
    """Load document using Docling (97.9% table accuracy)."""
    try:
        from docling.document_converter import DocumentConverter
        converter = DocumentConverter()
        result = converter.convert(str(file_path))
        return result.document.export_to_markdown()
    except Exception as e:
        return f"Error: {str(e)}"


def load_with_pymupdf(file_path: Path) -> str:
    """Fast fallback using pymupdf4llm."""
    try:
        import pymupdf4llm
        return pymupdf4llm.to_markdown(str(file_path))
    except Exception as e:
        return f"Error: {str(e)}"


def load_with_ocr(file_path: Path) -> tuple[str, List[str]]:
    """Load using EasyOCR for scanned documents."""
    try:
        suffix = file_path.suffix.lower()
        all_text = []
        text_by_page = []

        if suffix == '.pdf':
            from pdf2image import convert_from_path
            images = convert_from_path(str(file_path), dpi=200)

            for i, img in enumerate(images[:20], 1):
                temp_path = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                img.save(temp_path.name)
                result = extract_with_easyocr(temp_path.name)
                page_text = result['text']
                all_text.append(f"--- Page {i} ---\n{page_text}")
                text_by_page.append(page_text)
                os.unlink(temp_path.name)
        else:
            result = extract_with_easyocr(str(file_path))
            all_text.append(result['text'])
            text_by_page.append(result['text'])

        return '\n\n'.join(all_text), text_by_page

    except Exception as e:
        return f"Error: {str(e)}", []


def load_with_vision(file_path: Path) -> str:
    """Load using LLaVA vision model."""
    try:
        from pdf2image import convert_from_path
        import io

        suffix = file_path.suffix.lower()
        images_b64 = []

        if suffix == '.pdf':
            images = convert_from_path(str(file_path), dpi=150)
            for img in images[:5]:
                buffer = io.BytesIO()
                img.save(buffer, format='JPEG', quality=85)
                images_b64.append(base64.b64encode(buffer.getvalue()).decode())
        elif suffix in ['.jpg', '.jpeg', '.png', '.tiff', '.bmp']:
            with open(file_path, 'rb') as f:
                images_b64.append(base64.b64encode(f.read()).decode())
        else:
            return f"Vision not supported for {suffix}"

        llm = ChatOllama(
            model=VISION_MODEL,
            base_url=OLLAMA_BASE_URL,
            temperature=0
        )

        all_text = []
        for i, img_b64 in enumerate(images_b64, 1):
            message = HumanMessage(
                content=[
                    {"type": "text", "text": "Extract all text, tables, and structured information from this document. Format as markdown."},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}}
                ]
            )
            response = llm.invoke([message])
            all_text.append(f"--- Page {i} ---\n{response.content}")

        return '\n\n'.join(all_text)

    except Exception as e:
        return f"Error: {str(e)}"


# =============================================================================
# MAIN PROCESSOR CLASS
# =============================================================================

class DocumentProcessor:
    """Main document processing engine."""

    def __init__(self):
        self.output_dir = OUTPUT_DIR

    def process_document(
        self,
        file_path: str,
        method: str = "auto",
        extract_tables: bool = True,
        extract_signatures: bool = True,
        extract_key_values: bool = True,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> ExtractionResult:
        """
        Process a document and extract all structured data.

        Args:
            file_path: Path to the document
            method: Processing method (auto, docling, pymupdf, ocr, vision)
            extract_tables: Whether to extract tables
            extract_signatures: Whether to detect signatures
            extract_key_values: Whether to extract key-value pairs
            progress_callback: Optional callback for progress updates (percent, step_name)

        Returns:
            ExtractionResult with all extracted data
        """
        start_time = time.time()
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        def update_progress(percent: int, step: str):
            if progress_callback:
                progress_callback(percent, step)

        update_progress(5, "Initializing")

        suffix = file_path.suffix.lower()
        is_image = suffix in ['.jpg', '.jpeg', '.png', '.tiff', '.bmp', '.gif']
        is_pdf = suffix == '.pdf'

        # Determine method
        if method == "auto":
            if is_image:
                method = "ocr"
            else:
                method = "docling"

        update_progress(10, "Loading document")

        # Extract text
        text = ""
        text_by_page = []

        if method == "docling":
            text = load_with_docling(file_path)
            if text.startswith("Error:"):
                text = load_with_pymupdf(file_path)
        elif method == "pymupdf":
            text = load_with_pymupdf(file_path)
        elif method == "ocr":
            text, text_by_page = load_with_ocr(file_path)
        elif method == "vision":
            text = load_with_vision(file_path)

        update_progress(40, "Text extracted")

        # Convert to images for further processing
        images = []
        if is_pdf:
            from pdf2image import convert_from_path
            pil_images = convert_from_path(str(file_path), dpi=200)
            for img in pil_images[:20]:
                temp_path = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                img.save(temp_path.name)
                images.append(temp_path.name)
        elif is_image:
            images.append(str(file_path))

        # Extract tables
        all_tables = []
        if extract_tables and images:
            update_progress(50, "Extracting tables")
            for page_num, img_path in enumerate(images, 1):
                tables = extract_tables_with_img2table(img_path)
                for t in tables:
                    t.page = page_num
                all_tables.extend(tables)

        update_progress(65, "Tables extracted")

        # Detect signatures
        all_signatures = []
        review_items = []
        if extract_signatures and images:
            update_progress(70, "Detecting signatures")
            for page_num, img_path in enumerate(images, 1):
                sig_result = detect_signatures(img_path)
                for sig in sig_result['signatures']:
                    sig.page = page_num
                all_signatures.extend(sig_result['signatures'])
                review_items.extend(sig_result['human_review_items'])

        update_progress(80, "Signatures detected")

        # Extract key-values
        all_key_values = []
        if extract_key_values:
            update_progress(85, "Extracting key-value pairs")
            kv_pattern = r'([A-Za-z][A-Za-z\s]{2,30}):\s*([^\n:]{1,100})'
            matches = re.findall(kv_pattern, text)
            for key, value in matches:
                all_key_values.append(KeyValuePair(
                    key=key.strip(),
                    value=value.strip(),
                    confidence=0.8
                ))

        update_progress(90, "Key-values extracted")

        # Cleanup temp files
        if is_pdf:
            for img_path in images:
                try:
                    os.unlink(img_path)
                except:
                    pass

        # Calculate confidence
        confidence_scores = []
        if all_signatures:
            confidence_scores.extend([s.confidence for s in all_signatures])
        if all_key_values:
            confidence_scores.extend([kv.confidence for kv in all_key_values])
        overall_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0.7

        # Build result
        processing_time = time.time() - start_time

        result = ExtractionResult(
            document_source=str(file_path),
            document_type=self._detect_document_type(text),
            pages=len(images) if images else 1,
            processed_at=datetime.now().isoformat(),
            processing_time_seconds=round(processing_time, 2),
            text=text[:DOC_TEXT_LIMIT] if len(text) > DOC_TEXT_LIMIT else text,
            text_by_page=text_by_page if text_by_page else None,
            key_values=all_key_values,
            tables=all_tables,
            signatures=all_signatures,
            human_review_required=len(review_items) > 0,
            human_review_items=review_items,
            overall_confidence=round(overall_confidence, 3),
            warnings=[]
        )

        update_progress(100, "Complete")

        return result

    def _detect_document_type(self, text: str) -> Optional[str]:
        """Detect document type from content."""
        text_lower = text.lower()

        if any(word in text_lower for word in ['invoice', 'bill to', 'amount due', 'total due']):
            return "invoice"
        elif any(word in text_lower for word in ['contract', 'agreement', 'hereby agree', 'terms and conditions']):
            return "contract"
        elif any(word in text_lower for word in ['receipt', 'paid', 'transaction']):
            return "receipt"
        elif any(word in text_lower for word in ['resume', 'curriculum vitae', 'work experience', 'education']):
            return "resume"
        elif any(word in text_lower for word in ['form', 'application', 'please fill']):
            return "form"
        else:
            return None

    def save_result(self, result: ExtractionResult, output_path: Optional[str] = None) -> str:
        """Save extraction result to JSON file."""
        if output_path is None:
            source_name = Path(result.document_source).stem
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.output_dir / f"{source_name}_{timestamp}.json"
        else:
            output_path = Path(output_path)

        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, 'w') as f:
            json.dump(result.model_dump(), f, indent=2, default=str)

        return str(output_path)

    def export_to_csv(self, result: ExtractionResult, output_path: str) -> str:
        """Export key-values and tables to CSV."""
        import csv

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, 'w', newline='') as f:
            writer = csv.writer(f)

            # Key-values section
            writer.writerow(['KEY-VALUE PAIRS'])
            writer.writerow(['Key', 'Value', 'Confidence'])
            for kv in result.key_values:
                writer.writerow([kv.key, kv.value, kv.confidence])

            writer.writerow([])

            # Tables section
            for table in result.tables:
                writer.writerow([f'TABLE: {table.id}'])
                for row in table.rows:
                    writer.writerow(row)
                writer.writerow([])

            # Signatures section
            writer.writerow(['SIGNATURES'])
            writer.writerow(['ID', 'Status', 'Confidence', 'Page'])
            for sig in result.signatures:
                writer.writerow([sig.id, sig.status.value, sig.confidence, sig.page])

        return str(output_path)

    def export_to_excel(self, result: ExtractionResult, output_path: str) -> str:
        """Export to Excel with multiple sheets."""
        try:
            import openpyxl
            from openpyxl import Workbook

            wb = Workbook()

            # Key-values sheet
            ws_kv = wb.active
            ws_kv.title = "Key-Value Pairs"
            ws_kv.append(['Key', 'Value', 'Confidence', 'Page'])
            for kv in result.key_values:
                ws_kv.append([kv.key, kv.value, kv.confidence, kv.page])

            # Tables sheet
            ws_tables = wb.create_sheet("Tables")
            row_num = 1
            for table in result.tables:
                ws_tables.cell(row=row_num, column=1, value=f"Table: {table.id} (Page {table.page})")
                row_num += 1
                for row in table.rows:
                    for col_num, cell_value in enumerate(row, 1):
                        ws_tables.cell(row=row_num, column=col_num, value=cell_value)
                    row_num += 1
                row_num += 1

            # Signatures sheet
            ws_sig = wb.create_sheet("Signatures")
            ws_sig.append(['ID', 'Status', 'Confidence', 'Page', 'Location'])
            for sig in result.signatures:
                ws_sig.append([
                    sig.id, sig.status.value, sig.confidence, sig.page,
                    str(sig.location)
                ])

            # Summary sheet
            ws_summary = wb.create_sheet("Summary")
            ws_summary.append(['Document Source', result.document_source])
            ws_summary.append(['Document Type', result.document_type or 'Unknown'])
            ws_summary.append(['Pages', result.pages])
            ws_summary.append(['Processed At', result.processed_at])
            ws_summary.append(['Processing Time (s)', result.processing_time_seconds])
            ws_summary.append(['Key-Value Pairs', len(result.key_values)])
            ws_summary.append(['Tables', len(result.tables)])
            ws_summary.append(['Signatures', len(result.signatures)])
            ws_summary.append(['Human Review Required', result.human_review_required])
            ws_summary.append(['Overall Confidence', result.overall_confidence])

            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)

            return str(output_path)

        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")


# Global processor instance
processor = DocumentProcessor()
